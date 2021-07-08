import jira.client
from jira.client import JIRA
from openpyxl import Workbook

headerslogin = {
    "Content-Type": "application/json",
    "Authorization": "Basic ZTE3NDE1OjZhNjZhNjZhNj=="
}

jira_options = {'server': 'https://jira.g2-networks.net/',
                'headers': headerslogin}
jira = JIRA(options=jira_options)

key_list = []
summary_list = []

# Add additional lists for fields here
description_list = []

# Get All assigned Issues of a particular user
# issues_in_project = jira.search_issues('assignee=e17415')
# Get All assigned Issues created today of a particular user
# issues_in_project = jira.search_issues('assignee=e17415 and created > startOfDay(-0d)')
# Get All assigned STAGE Issues of a particular user
# issues_in_project = jira.search_issues('assignee=e17415 and labels in (configurazionegiochistage)')
# Get All assigned LIVE Issues of a particular user
issues_in_project = jira.search_issues('assignee=e17415 and labels in (configurazionegiochilive)')
# Get All assigned LIVE OPNED Issues of a particular user
# issues_in_project = jira.search_issues('assignee=e17415 and Status in (1) and labels in (configurazionegiochilive)')
# Get All assigned STAGE OPNED Issues of a particular user
# issues_in_project = jira.search_issues('assignee=e17415 and Status in (1) and labels in (configurazionegiochistage)')
# Get All assigned STAGE IN PROGRESS Issues of a particular user
# issues_in_project = jira.search_issues('assignee=e17415 and Status in (3) and labels in (configurazionegiochistage)')
# Get All assigned LIVE IN PROGRESS Issues of a particular user
# issues_in_project = jira.search_issues('assignee=e17415 and Status in (3) and labels in (configurazionegiochilive)')

for issue in issues_in_project:
    key_list.append(issue.key)
    summary_list.append(issue.fields.summary)
    # Add additional fields returned here
    description_list.append(issue.fields.description)

wb = Workbook()
ws = wb.active
key_row = 1
summary_row = 1

description_row = 1

start_column = 1

for key in key_list:
    ws.cell(row=key_row, column=start_column).value = key
key_row += 1

for summary in summary_list:
    ws.cell(row=summary_row, column=start_column + 1).value = summary
summary_row += 1

# add additional fields here

for description in description_list:
    ws.cell(row=description_row, column=start_column + 2).value = description
description_row += 1

wb.save("jira-report1.xlsx")
